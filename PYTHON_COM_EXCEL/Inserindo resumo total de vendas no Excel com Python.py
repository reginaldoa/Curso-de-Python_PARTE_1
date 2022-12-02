from openpyxl import load_workbook
import os


Caminho_arquivo = 'C:\\Users\\viola\\OneDrive\\Documentos\\CURSOS TECNOLOGIA\\UDEMY\\pythonProject\\PYTHON_COM_EXCEL\\Vendedores.xlsx'
planilha_que_abrimos_com_python = load_workbook(filename = Caminho_arquivo)

#Selecionando a planilha e abrindo a mesma
sheet_selecionada = planilha_que_abrimos_com_python ['Vendas']

somaAmandaMartins = 0
somaElianeMoreira = 0
somaLeonardoAlmeida = 0
somaNicolasPereira = 0

for i in range(2, len(sheet_selecionada['C']) +1):
    if sheet_selecionada['A%s' % i].value == "Amanda Martins":
        somaAmandaMartins = somaAmandaMartins + sheet_selecionada['C%s'% i].value

    elif sheet_selecionada['A%s' % i].value == "Eliane Moreira":
        somaElianeMoreira = somaElianeMoreira  + sheet_selecionada['C%s'% i].value

    elif sheet_selecionada['A%s' % i].value == "Leonardo Almeida":
        somaLeonardoAlmeida = somaLeonardoAlmeida  + sheet_selecionada['C%s'% i].value

    elif sheet_selecionada['A%s' % i].value == "Nicolas Pereira":
        somaNicolasPereira = somaNicolasPereira  + sheet_selecionada['C%s' % i].value


sheet_resumo = planilha_que_abrimos_com_python.create_sheet("Resumo")
sheet_selecionada = planilha_que_abrimos_com_python['Resumo']

#Título
sheet_selecionada["A1"] = "Vendedor"
sheet_selecionada["B1"] = "Vendas"

#Preenche os dados do vendedor
sheet_selecionada["A2"] = "Amanda Martins"
sheet_selecionada["B2"] = somaAmandaMartins

#Preenche os dados do vendedor
sheet_selecionada["A3"] = "Eliane Moreira"
sheet_selecionada["B3"] = somaElianeMoreira


#Preenche os dados do vendedor
sheet_selecionada["A5"] = "Leonardo Almeida"
sheet_selecionada["B5"] = somaLeonardoAlmeida


#Preenche os dados do vendedor
sheet_selecionada["A6"] = "Nicolas Pereira"
sheet_selecionada["B6"] = somaNicolasPereira



#Salvando as alterações
planilha_que_abrimos_com_python.save(filename=Caminho_arquivo )
print("Processo encerrado")


#Abrindo o arquivo
os.startfile(Caminho_arquivo)